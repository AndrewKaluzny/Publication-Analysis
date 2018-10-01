import nltk
from pathlib import Path
from nltk.tokenize import regexp_tokenize, sent_tokenize
from openpyxl import Workbook, load_workbook


filename = input("Type in the name of the text file (include the '.txt'):   ")
file = open(filename, "r+", encoding='utf-8-sig')
excelfile = filename[:-6] + '.xlsx'
my_file = Path(excelfile)
if my_file.is_file():
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
else:
    wb = Workbook(excelfile)
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws['A1'] = "# of Sentences: "
    ws['A2'] = "# of Words: "
    ws['A3'] = "# of Unique Words: "
    ws['A4'] = "# of Characters: "
    ws['A5'] = "Average word length: "
    ws['A6'] = "Average sentence length: "
    ws['A7'] = "Noun %: "
    ws['A8'] = "Verb %: "
    ws['A9'] = "Adjective %: "
    ws['A10'] = "Adverb %: "
    ws['A11'] = "Pronoun %: "
    ws['A12'] = "Preposition %: "
    ws['A13'] = "Coordinating Conjunction %: "
    ws['A14'] = "Miscellaneous Part of Speech %: "
text = file.read()


class article_base:
    def __init__(self, name, text):

# CONSTRUCTOR: Tokenize Sentences, Expressions
# *******************************************************************************************
        self.name = name
        text = text.lower()
        self.sent_t = sent_tokenize(text)
        self.word_t = regexp_tokenize(text, r'\w+')
        self.nword = len(self.word_t)

# CONSTRUCTOR: Unique Words (+ sort), Avg Sent/Word Length
# *******************************************************************************************
        self.unique_words = {}
        self.nsent = len(self.sent_t)
        for word in self.word_t:
            if word not in self.unique_words:
                self.unique_words[word] = 1
            else:
                self.unique_words[word] += 1
        count = 0;
        self.nuword = len(self.unique_words)
        self.unique_words_sorted = [None] * self.nuword
        for word in self.unique_words:
            self.unique_words_sorted[count] = (word, self.unique_words[word])
            count += 1
        self.unique_words_sorted = sorted(self.unique_words_sorted, key = lambda tuple: tuple[1])
        self.unique_words_sorted.reverse()
        self.char_count = 0
        for word in self.word_t:
            self.char_count += len(word)
        self.avg_word_length = round((float(self.char_count)/float(self.nword)), 2)
        self.avg_sent_length = round((float(self.nword)/float(self.nsent)), 2)

# CONSTRUCTOR: Parts of Speech
# *******************************************************************************************
        noun = ['NN', 'NNS', 'NNP', 'NNPS', 'POS']
        verb = ['VB', 'VBD', 'VBG', 'VBN', 'VBP', 'VBZ', 'RP', 'MD']
        adj = ['JJ', 'JJR', 'JJS', 'PDT', 'WDT', 'WP$']
        adv = ['RB', 'RBR', 'RBS']
        pro = ['PRP', 'PRP$', 'WP']
        prep = ['IN', 'TO', 'WRB']
        cc = ['CC']
        misc = ['CD', 'DT', 'EX', 'FW', 'LS', 'UH']
        self.pos_list = nltk.pos_tag(self.word_t)
        self.noun_count = 0
        self.verb_count = 0
        self.adj_count = 0
        self.adv_count = 0
        self.pro_count = 0
        self.prep_count = 0
        self.cc_count = 0
        self.misc_count = 0
        for word in self.pos_list:
            if word[1] in noun:
                self.noun_count += 1
            elif word[1] in verb:
                self.verb_count += 1
            elif word[1] in adj:
                self.adj_count += 1
            elif word[1] in adv:
                self.adv_count += 1
            elif word[1] in pro:
                self.pro_count += 1
            elif word[1] in prep:
                self.prep_count += 1
            elif word[1] in cc:
                self.cc_count += 1
            elif word[1] in misc:
                self.misc_count += 1
        self.noun_percent = round((100*float(self.noun_count)/float(self.nword)), 2)
        self.verb_percent = round((100*float(self.verb_count)/float(self.nword)), 2)
        self.adj_percent = round((100*float(self.adj_count)/float(self.nword)), 2)
        self.adv_percent = round((100*float(self.adv_count)/float(self.nword)), 2)
        self.pro_percent = round((100*float(self.pro_count)/float(self.nword)), 2)
        self.prep_percent = round((100*float(self.prep_count)/float(self.nword)), 2)
        self.cc_percent = round((100*float(self.cc_count)/float(self.nword)), 2)
        self.misc_percent = round((100*float(self.misc_count)/float(self.nword)), 2)

# FUNCTIONS: Print Variables
# *******************************************************************************************
    def print_num_sent(self):
        print("# of Sentences: " + str(self.nsent))
        ws.cell(row=1, column=ws.max_column+1).value = self.nsent

    def print_num_word(self):
        print("# of Words: " + str(self.nword))
        ws.cell(row=2, column=ws.max_column).value = self.nword

    def print_num_uword(self):
        print("# of Unique Words: " + str(self.nuword))
        ws.cell(row=3, column=ws.max_column).value = self.nuword

    def print_char_count(self):
        print("# of Characters: " + str(self.char_count))
        ws.cell(row=4, column=ws.max_column).value = self.char_count

    def print_avg_word_length(self):
        print("Average word length: " + str("{0:.2f}".format(self.avg_word_length)) + " characters")
        ws.cell(row=5, column=ws.max_column).value = self.avg_word_length

    def print_avg_sent_length(self):
        print("Average sentence length: " + str("{0:.2f}".format(self.avg_sent_length)) + " words")
        ws.cell(row=6, column=ws.max_column).value = self.avg_sent_length

    def print_pos_count(self):
        print("Noun %: " + str("{0:.2f}".format(self.noun_percent)))
        ws.cell(row=7, column=ws.max_column).value = self.noun_percent
        print("Verb %: " + str("{0:.2f}".format(self.verb_percent)))
        ws.cell(row=8, column=ws.max_column).value = self.verb_percent
        print("Adjective %: " + str("{0:.2f}".format(self.adj_percent)))
        ws.cell(row=9, column=ws.max_column).value = self.adj_percent
        print("Adverb %: " + str("{0:.2f}".format(self.adv_percent)))
        ws.cell(row=10, column=ws.max_column).value = self.adv_percent
        print("Pronoun %: " + str("{0:.2f}".format(self.pro_percent)))
        ws.cell(row=11, column=ws.max_column).value = self.pro_percent
        print("Preposition %: " + str("{0:.2f}".format(self.prep_percent)))
        ws.cell(row=12, column=ws.max_column).value = self.prep_percent
        print("Coordinating Conjunction %: " + str("{0:.2f}".format(self.cc_percent)))
        ws.cell(row=13, column=ws.max_column).value = self.cc_percent
        print("Miscellaneous Part of Speech %: " + str("{0:.2f}".format(self.misc_percent)))
        ws.cell(row=14, column=ws.max_column).value = self.misc_percent

# FUNCTIONS: List Variables
# *******************************************************************************************
    def sent_list(self):
        print("\n   List of Sentences: ")
        for sent in self.sent_t:
            count = 0
            for word in regexp_tokenize(sent, r'\w+'):
                count += 1
            print(str(count) + " words: " + sent)

    def word_list(self):
        print("\n   List of Words: ")
        for word in self.word_t:
            print(word)

    def uword_list(self):
        print("\n   List of Unique Words: ")
        for tup in self.unique_words_sorted:
            print("%10s %15s %10s" % (str(self.unique_words_sorted.index(tup)+1), str(tup[0]), str(tup[1])))

# FUNCTIONS: 2nd order Print (combinations)
# *******************************************************************************************
    def full_stat_print(self):
        self.print_num_sent()
        self.print_num_word()
        self.print_num_uword()
        self.print_char_count()
        self.print_avg_word_length()
        self.print_avg_sent_length()
        self.print_pos_count()
        self.sent_list()
        self.uword_list()

    def quick_stat_print(self):
        self.print_num_sent()
        self.print_num_word()
        self.print_num_uword()
        self.print_char_count()
        self.print_avg_sent_length()
        self.print_avg_word_length()
        self.print_pos_count()


test = article_base("base_one", text)
ans = ''
while ans != 'Q' and ans != 'q' and ans != 'F' and ans != 'f':
    ans = input("Would you like a (q)uick stat print, or the (f)ull works?    ")
if ans == 'Q' or ans == 'q':
    test.quick_stat_print()
else:
    test.full_stat_print()

wb.save(excelfile)